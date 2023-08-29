import json
import os
import openpyxl

current_dir = os.path.dirname(os.path.abspath(__file__))
data_dir = os.path.join(current_dir, "data")
skill_path = os.path.join(data_dir, "skill.json")
task_path = os.path.join(data_dir, "task.json")


async def skill_to_json() -> None:
    """Конвертирует данные из листа 'Skill' Excel-файла в JSON-формат."""
    workbook = openpyxl.load_workbook('file.xlsx')
    sheet = workbook['Skill']

    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if isinstance(row[1], float):
            if_passed_value = int(row[1])
        else:
            if_passed_value = row[1]

        if isinstance(row[2], float):
            if_failed_value = int(row[2])
        else:
            if_failed_value = row[2]

        row_data = {
            'skill_id': int(row[0]),
            'if_passed': if_passed_value,
            'if_failed': if_failed_value,
            'task_id': list(map(int, row[3].split(';')))
        }
        data.append(row_data)

    json_data = json.dumps(data, indent=2, ensure_ascii=False)

    with open(skill_path, 'w', encoding='utf-8') as json_file:
        json_file.write(json_data)


async def task_to_json() -> None:
    """Конвертирует данные из листа 'Task' Excel-файла в JSON-формат."""
    workbook = openpyxl.load_workbook('file.xlsx')
    sheet = workbook['Task']

    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_data = {
            'task_id': int(row[0]),
            'question': row[1],
            'answer': int(row[2]),
        }
        data.append(row_data)

    json_data = json.dumps(data, indent=2, ensure_ascii=False)

    with open(task_path, 'w', encoding='utf-8') as json_file:
        json_file.write(json_data)
